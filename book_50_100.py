#!/usr/bin/python
import requests
import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
# 2500/3:22pm
#34Paginas son 1020 registros
#4052
# maxpag 334
id_book = 0
start_range = 50
continue_next = False
title_book = ''
name_file = 'book_50_100.xlsx'
wb = load_workbook(name_file, read_only=True)
ws = wb.worksheets[0]
max_row = ws.max_row
if max_row > 1:
    print('Loading row {}'.format(max_row))

    if ws.cell(row=max_row, column=1).value != 0 and ws.cell(row=max_row, column=1).value != None:
        id_book = ws.cell(row=max_row, column=1).value
        title_book = ws.cell(row=max_row, column=2).value
        start_range = ws.cell(row=max_row, column=15).value

print('Finish loading: {} {} {}'.format(id_book, start_range, title_book))
book_links = []
#150 - 200
for itera in range(start_range, 100):
    url = 'https://www.bookdepository.com/category/2/Art-Photography/browse/viewmode/all?page={}'.format(itera)
    req = requests.get(url)
    soup = BeautifulSoup(req.text, "lxml")

    book_links = [book.a.attrs['href'] for book in soup.find_all('h3', 'title')]

    for i, book in enumerate(book_links):

        if title_book == '':
            continue_next = True
    
        if continue_next:
            id_book = id_book + 1
    
        name_book = ''  #Obligatorio o no guardar registro
        author_book = ''  #Obligatorio o no guardar registro
        publisher_book = '' #Obligatorio o no guardar registro
        datepublished_book = ''
        language_book = 'Not Available'
        number_of_pages_book = ''
        isbn_book = ''
        description_book = 'Not Available'
        format_book = 'Not Available'
        image_book = ''
        url_book = ''
        pagination_book = itera
        subcategoria_1_book = ''
        subcategoria_2_book = ''
        subcategoria_3_book = ''
        subcategoria_4_book = ''
        subcategoria_5_book = ''
        subcategoria_6_book = ''
        subcategoria_7_book = ''
        subcategoria_8_book = ''
        subcategoria_9_book = ''
        subcategoria_10_book = ''
        price_book = '$'
        dimension_book = 'Not Available'
        weight_book = 'Not Available'
        imprint_book = 'Not Available'
        published_in_book = 'Not Available'

        url_book = 'https://www.bookdepository.com{}'.format(book)
        req = requests.get(url_book)
        soup = BeautifulSoup(req.text, "lxml")

        try: 
            name_book = soup.find('h1').text.strip()
        except: 
            pass
        
        try: 
            author_book = soup.find('span', itemprop="author").text.strip()
        except: 
            pass
        
        try: 
            publisher_book = soup.find('span', itemprop="publisher").text.strip()
        except: 
            pass
        
        try: 
            datepublished_book = soup.find('span', itemprop="datePublished").text.strip()[-4:]
        except: 
            pass
        
        try: 
            label = soup.find('label', text=["Language"])
            language_book = label.find_next_sibling('span').text.strip()
        except: 
            pass

        # try: 
        #     number_of_pages_book = soup.find('span', itemprop="numberOfPages").text.strip().replace(' pages','')
        # except: 
        #     pass

        try: 
            isbn_book = soup.find('span', itemprop="isbn").text.strip()

            if not isbn_book:
                label = soup.find('label', text=["ISBN10"])
                isbn_book = label.find_next_sibling('span').text.strip()
        except: 
            pass
        
        try: 
            description_book = soup.find('div', itemprop="description").text.replace('show more\n','').replace('\n','')
        except: 
            pass

        try:
            label = soup.find('label', text=["Format"])
            formats = label.find_next_sibling('span').text.strip().replace('\n','')
            format_book = formats
            
            if '|' in formats:
                formats = formats.split('|')
                format_book = formats[0].strip()
                number_of_pages_book = formats[1].replace('pages','').strip()
        except: 
            pass
        
        try: 
            image_book = soup.find('img', 'book-img').attrs['src']
        except: 
            pass
        
        try: 
            categories_ol = soup.find('ol', 'breadcrumb')
            for index, li in enumerate(categories_ol.find_all('li')):
                if index !=0:
                    if index == 1:
                        subcategoria_1_book = li.text.strip()
                    if index == 2:
                        subcategoria_2_book = li.text.strip()
                    if index == 3:
                        subcategoria_3_book = li.text.strip()
                    if index == 4:
                        subcategoria_4_book = li.text.strip()
                    if index == 5:
                        subcategoria_5_book = li.text.strip()
                    if index == 6:
                        subcategoria_6_book = li.text.strip()
                    if index == 7:
                        subcategoria_7_book = li.text.strip()
                    if index == 8:
                        subcategoria_8_book = li.text.strip()
                    if index == 9:
                        subcategoria_9_book = li.text.strip()
                    if index == 10:
                        subcategoria_10_book = li.text.strip()
        except: 
            pass

        try: 
            price_book = soup.find('span', 'sale-price').text.strip()
        except: 
            pass

        try: 
            label = soup.find('label', text=["Dimensions"])
            dimensions = label.find_next_sibling('span').text.replace('\n','').replace(' ','').strip()
            dimension_book = dimensions
            
            if '|' in dimensions:
                dimensions = dimensions.split('|')
                dimension_book = dimensions[0]
                weight_book = dimensions[1]
        except: 
            pass

        try: 
            label = soup.find('label', text=["Imprint"])
            imprint_book = label.find_next_sibling('span').text.strip()
        except: 
            pass

        try: 
            label = soup.find('label', text=["Publication City/Country"])
            published_in_book = label.find_next_sibling('span').text.strip()
        except: 
            pass

        if continue_next:
            row = [id_book, name_book, author_book, ' ', publisher_book, datepublished_book, language_book, number_of_pages_book, 'Art & Photography', isbn_book, description_book, format_book, image_book, url_book, pagination_book, subcategoria_1_book, subcategoria_2_book, subcategoria_3_book, subcategoria_4_book, subcategoria_5_book, subcategoria_6_book, subcategoria_7_book, subcategoria_8_book, subcategoria_9_book, subcategoria_10_book, price_book, dimension_book, weight_book, imprint_book, published_in_book]

            wb = load_workbook(name_file)
            # # Select First Worksheet
            ws = wb.worksheets[0]
            ws.append(row)

            wb.save(name_file)

        if title_book != '' and title_book==name_book:
            continue_next = True

        print(id_book, pagination_book, name_book)
    # input(">_ ")
